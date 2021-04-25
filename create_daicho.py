import openpyxl
import glob
import csv
from pprint import pprint
import re

# aaaaaa

new_cols_dict = {
    "日付": 1,
    "決済方法": 3,
    "注文時間": 4,
    "リピーターフラグ":5,
    "注文者": 6,
    "注文者会社名":7,
    "商品名":8,
    "個数":9,
    "商品価格":10,
    "送料":11,
    "手数料":12,
    "消費税":13,
    "注文金額":14,
    "処理状況":15,
    "方法":16,
    "出荷日":17,
    "担当":18,
    "件数": 2
}

dates_dict = {}

payments_dict = {
    "代金引き換え":"D",
    "NP掛け払い(FREX B2B 後払)":"NK",
    "NP後払い(請求書後払い)":"NP"
}

# workbooks = glob.glob("excel/*")
# for i in range(0, len(workbooks)):
#     if "csv" in workbooks[i]:
#         wb = openpyxl.load_workbook(workbooks[i])
#         ws = wb["Sheet1"]

wb_daicho = openpyxl.load_workbook('daicho01.xlsx')
ws_daicho = wb_daicho["受注管理表"]

with open("excel/honten_dummy20210417.csv", "r", encoding="shift-jis") as f:
    reader = csv.DictReader(f)
    # csv_list = [row for row in reader]
    # print(csv_list[1]["日付"])
    for row in reader:
        for key, new_col in new_cols_dict.items():
            if key == "日付":
                # 2021-04-12を04/12に置換
                cut_date = re.sub(r"2021-0|2021-", "", row[key])
                replace_date = cut_date.replace("-", "/")
                ws_daicho.cell(row=reader.line_num+1, column=new_col, value=replace_date)

                # 日付をカウント
                date = str(row[key])
                if date in dates_dict:
                    count = dates_dict[date] + 1
                else:
                    count = 1
                dates_dict[date] = count
                continue

            if key == "件数":
                # 一つ前のデータ
                ueno_date = ws_daicho.cell(reader.line_num, 1).value
                ueno_time = ws_daicho.cell(reader.line_num, 4).value
                ueno_name = ws_daicho.cell(reader.line_num, 6).value
                # 記入する行のデータ
                now_row_date = ws_daicho.cell(reader.line_num+1, 1).value
                now_row_time = ws_daicho.cell(reader.line_num+1, 4).value
                now_row_name = ws_daicho.cell(reader.line_num+1, 6).value
                
                if now_row_date != ueno_date:
                    i = 1
                elif now_row_time != ueno_time:
                    i = 1 + ws_daicho.cell(reader.line_num, new_col).value
                elif now_row_name == ueno_name:
                    i = ws_daicho.cell(reader.line_num, new_col).value

                ws_daicho.cell(row=reader.line_num+1, column=new_col, value=i)
                continue

            if key == "決済方法":
                payment = str(row[key])
                ws_daicho.cell(row=reader.line_num+1, column=new_col, value=payments_dict[payment])
                continue

            if key =="リピーターフラグ":
                if row[key] == "リピーター":
                    ws_daicho.cell(row=reader.line_num+1, column=new_col, value="R")
                else:
                    ws_daicho.cell(row=reader.line_num+1, column=new_col, value="N")
                continue
            
            if key == "手数料" or key == "方法" or key == "出荷日" or key == "担当":
                continue

            ws_daicho.cell(row=reader.line_num+1, column=new_col, value=row[key])

print(dates_dict)
wb_daicho.save("daicho04.xlsx")


